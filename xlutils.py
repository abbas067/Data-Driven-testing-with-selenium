import openpyxl 

def getRowCount(file,sheetName): 

    workbook=openpyxl.load_workbook(file) 

    sheet=workbook.get_sheet_by_name(sheetName) 

    return (sheet.max_row) 

def getColsCount(file,sheetName): 

    workbook=openpyxl.load_workbook(file) 

    sheet=workbook.get_sheet_by_name(sheetName) 

    return (sheet.max_column) 

def readData(file,sheetName,rowno,columnno): 

    workbook = openpyxl.load_workbook(file) 

    sheet = workbook.get_sheet_by_name(sheetName) 

    return sheet.cell(row=rowno,column=columnno).value 

def writeData(file,sheetName,rowno,columnno,data): 

    workbook = openpyxl.load_workbook(file) 

    sheet = workbook.get_sheet_by_name(sheetName) 

    sheet.cell(row=rowno,column=columnno).value=data 

    workbook.save(file) 
